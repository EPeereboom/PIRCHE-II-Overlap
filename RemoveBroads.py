import argparse
from openpyxl import load_workbook

def readSplitToBroad(fileName=None):
    print('Loading serology from ' + str(fileName))
    splitToBroad = {}
    with open(fileName, "r") as bsfile:
        for row in bsfile:
            if not row.startswith("#"):
                cols = row.split(";")
                locus = cols[0]
                broad = cols[1]
                splits = cols[2].split("/")

                for split in splits:
                    splitToBroad[locus + split] = locus + broad
    return splitToBroad

def convertTyping(cell=None, splitToBroad=None):
    cellValue = cell.value
    allTypes = [str(typing).strip() for typing in str(cellValue).split(' ')]

    typesBroadsRemoved = allTypes.copy()
    for typing in allTypes:
        if typing in splitToBroad.keys():
            #print('This is a split:' + str(typing))
            # If the broad is already there, skip it.
            broad = splitToBroad[typing]
            if broad in typesBroadsRemoved:
                typesBroadsRemoved.remove(broad)
                print('\n' + str(cell.column_letter) + str(cell.row) + ': Broad ' + str(broad) + ' and split ' + typing + ' both exist. Removed the broad. \nBefore: '+ str(allTypes) + '\nAfter: ' + str(typesBroadsRemoved) + '')
            else:
                pass
                #print(str(cell.column_letter) + str(cell.col_idx) + ' Split ' + typing + ' does not have its broad ' + str(broad) + '. ('+ str(allTypes) + ')->(' + str(typesBroadsRemoved) + ')')

        else:
            pass
            #print(str(cell.column_letter) + str(cell.col_idx) + ' Not a split:' + str(typing))

    return ' '.join(typesBroadsRemoved)

def convertFile(args=None):
    splitToBroad=readSplitToBroad(fileName=args.serology)

    print('Reading (only the active sheet) file ' + str(args.input))
    workBook = load_workbook(args.input)
    workSheet = workBook.active

    columnLetters = args.columns.split(',')

    # Loop rows (skip header..)
    if args.headers:
        begin=2
    else:
        begin=1
    for row in workSheet.iter_rows(min_row=begin):
        for cell in row:
            if cell.column_letter in columnLetters:
                cell.value = convertTyping(cell=cell, splitToBroad=splitToBroad)

    # write the altered file
    outputFileName = args.input+'.fixed.xlsx'
    workBook.save(outputFileName)
    print('Done, output written to ' + outputFileName)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-H", "--headers", help="file contains header line", action="store_true")
    parser.add_argument("-i", "--input", help="immunizers input file", required=True)
    parser.add_argument("-s", "--serology", help="locate the rel_ser_ser.txt file from imgt", required=True)
    parser.add_argument("-f", "--columns", help="column letters to analyze, comma separated", required=True)

    args = parser.parse_args()

    verbose = args.verbose
    if verbose:
        print("verbose mode active")

    convertFile(args=args)
